#SECAO10
#Aula 62 e 63
#MANIPULANDO EXCEL
from openpyxl import load_workbook
import os

wb = load_workbook(filename='nomes.xlsx')

planilha = wb['Planilha1']

somaIdade= 0

for i in range(2, 5):
    print('%s tem %s anos.'% (planilha['A%s' % i].value, planilha['B%s'% i].value))
    somaIdade += int(planilha['B%s' % i].value)

planilha['B5'] = somaIdade

#add dados a lista
planilha['A7'] = 'ALUNOS'
#mesclar as celulas
planilha.merge_cells('A7:B7')
#desmeclar as celulas
planilha.unmerge_cells('A7:B7')

wb.save('nomes.xlsx')

os.startfile('nomes.xlsx')

#Aula 64
#Inserindo imagens

from openpyxl import Workbook
from openpyxl.drawing.image import Image

wb = Workbook()
ws = wb.active

ws.merge_cells('A1:D1')

ws['A1'] = 'Você vai ver uma imagem abaixo'

img = Image('BIDAO.jfif')

ws.add_image(img, 'A2')

wb.save('bida.xlsx')

os.startfile('bida.xlsx')


#Aula 66
#deletando linhas e coluna

arquivo = 'bookAula66.xlsx'
wb = load_workbook(arquivo)

ws = wb['Data']

#deleta linha(qual linha)
# ws.delete_rows(1)

#deleta coluna(qual coluna)
# ws.delete_cols(1)

wb.save(arquivo)

os.startfile(arquivo)

#Aula 67
#movendo celulas

arquivo = 'bookAula67.xlsx'

wb = load_workbook(arquivo)

ws = wb['Data']

#mover uma sequencia de ceulas
#seleciona a planilha, move selecionando quais linhas, subtrai -1 linha
ws.move_range('A2:Z2', rows=-1)

#move f8 duas colunas para a direita
ws.move_range('F8:F8', cols=2)

#move f8 duas colunas para a direita
ws.move_range('F10:F10', cols=-2)

#movendo C para cima e esquerda
ws.move_range('C13:E15', cols=-2, rows=-1)

wb.save(arquivo)
os.startfile(arquivo)























