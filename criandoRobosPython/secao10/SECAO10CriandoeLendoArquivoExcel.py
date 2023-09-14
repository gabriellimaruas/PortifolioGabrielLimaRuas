#SECAO 10 EXCEL COM OPENPYXL
#CRIANDO ARQUIVO EXCEL

from openpyxl import Workbook

workbook = Workbook()

arquivo = 'D:\\ProjetosPycharm\\criandoRobosPython\\secao10\\primeiroArquivo.xlsx'

ws1 = workbook.active
ws1.title = 'Planilha 1'

dados = [
    ['Ano', 'Lucro', 'Custos'],
    [2015, '30%', '40'],
    [2016, '50%', '40'],
    [2017, '70%', '40']
]

#for para colocar os dados no arquivo excel
for linha in dados:
    ws1.append(linha)

workbook.save(arquivo)


#LENDO DADOS DE UM ARQUIVO

from openpyxl import load_workbook

workbook = load_workbook('/secao10/nomes.xlsx')

planilha = workbook['Planilha1']


#usando range para percorrer um conjunto de numeros
for i in range(2,6):
    print('%s tem %s anos' % (planilha['A%s' % i].value, planilha['B%s' % i].value))
    #passo a passo do codigo de cima
    #print('Gabriel tem 23 anos')
    #print('%s tem %s anos' % (planilha['A2'].value, planilha['B2'].value))
    #print('%s tem %s anos' % (planilha['A%s' % i].value, planilha['B%s' % i].value))


