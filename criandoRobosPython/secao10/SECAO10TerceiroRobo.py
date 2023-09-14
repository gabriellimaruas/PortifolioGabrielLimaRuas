#SECAO 10
#TERCEIRO ROBO
from openpyxl import load_workbook, Workbook
import os
import time

#Lista de arquivos para ser consolidada
listaArquivos = ['CustosAutom', 'PopulacaoPOA', 'SuperMercado']

print('Iniciando nosso robo...')
time.sleep(1)
#novo arquivo
wb = Workbook()
nomeArquivoFinal = 'resultado.xlsx'

for nomeArquivo in listaArquivos:
    print('Lendo arquivo %s...' % nomeArquivo)
    time.sleep(1)
    arquivo = load_workbook('%s.xlsx' % nomeArquivo)
    sheet = arquivo[nomeArquivo]
    maxLinha = sheet.max_row
    maxColunas = sheet.max_column

    ws = wb.create_sheet(nomeArquivo)

    #passar os dados de um arquivo para o outro
    for i in range(1, maxLinha +1):
        for j in range(1, maxColunas +1):

            c = sheet.cell(row=i, column=j)

            ws.cell(row=i, column=j).value = c.value

print('Criando arquivo final')
time.sleep(1)
wb.remove(wb['Sheet'])
wb.save(nomeArquivoFinal)
os.startfile(nomeArquivoFinal)
